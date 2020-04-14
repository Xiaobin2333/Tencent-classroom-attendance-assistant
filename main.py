'''
腾讯课堂考勤助手：一款自动化考勤数据处理工具 V1.0.0 By.King
博客地址:https://www.mrchung.cn/archives/53/
项目地址:https://github.com/Xiaobin2333/Tencent-classroom-attendance-assistant
'''
import openpyxl
import xlrd
import time
import json
import re
import os

def read_config():
    try:
        config_file = open('./config.json', encoding='utf-8')
        config = json.load(config_file)
    except:
        print("配置文件打开失败!")
    try:
        global name_x, name_y, txkt_start_x, txkt_duration_y, txkt_name_y, class_y, data_path, class_path, min_class, min_num
        name_x, name_y, txkt_start_x, txkt_duration_y, txkt_name_y, class_y, data_path, class_path, min_class, min_num = \
        config['name_x'], config['name_y'], config['txkt_start_x'], config['txkt_duration_y'], config['txkt_name_y'], config[
            'class_y'], config['data_path'], config['class_path'], config['min_class'], config['min_num']
        return True
    except:
        print("配置文件读取出错!")
        time.sleep(3)
        exit()

def get_excel():  #获取全部考勤表信息
    data_excel = []
    dir1 = os.listdir(data_path)
    for dir2 in dir1:
        path = data_path + '/' + dir2
        name1 = os.listdir(path)
        for name2 in name1:
            excel_date = name2[:10]
            excel_path = path + '/' + name2
            excel_subject = dir2
            data_excel.append([excel_date, excel_path, excel_subject])
    return (data_excel)

def get_name():  #获取本班名单
    data_name = []
    data = xlrd.open_workbook(class_path)
    excel_class = data.sheet_names()
    table = data.sheet_by_name(class_name)
    nrows = table.nrows
    for excel_x in range(name_x, nrows):
        name = table.cell(excel_x, name_y).value
        data_name.append(name)
    return (data_name)

def get_class_data(excel_path, data_name):  #获取单个考勤表数据
    data_list = []
    data = xlrd.open_workbook(excel_path)
    table = data.sheet_by_name('数据导出')
    nrows = table.nrows
    for excel_x in range(txkt_start_x, nrows):
        excel_name = table.cell(excel_x, txkt_name_y).value
        for name in data_name:
            if bool(re.search(name, excel_name)) == True:
                duration = table.cell(excel_x, txkt_duration_y).value
                duration_num = re.findall(r'\d+', duration)
                if duration == '不足一分钟':
                    duration_num = 0
                else:
                    duration_num = int(duration_num[0])
                data_list.append([name, duration_num])
    #通过名字去重并叠加时间(为了方便index分成两个list)
    name_list = []
    duration_list = []
    for j in range(0, len(data_list)):
        list_name = data_list[j][0]
        list_duration = data_list[j][1]
        if list_name not in name_list:
            name_list.append(list_name)
            duration_list.append(list_duration)
        else:
            p = name_list.index(list_name)
            duration_list[p] += list_duration
    new_data_list = []
    for k in range(0, len(name_list)):
        new_data_list.append([name_list[k], duration_list[k]])
    return (new_data_list)

def get_class(data_excel, data_name):  #获取全部考勤表数据
    data_class = []
    for excel in data_excel:
        excel_path = excel[1]
        excel_date = excel[0]
        excel_subject = excel[2]
        class_data = get_class_data(excel_path, data_name)
        data_class.append([excel_date, excel_subject, class_data])
    return (data_class)

def get_noclass_data(data_name, class_data):  #获取单个考勤表缺勤数据
    noclass_data = []
    name_list = []
    duration_list = []
    for data_list in class_data:
        name_list.append(data_list[0])
        duration_list.append(data_list[1])
    for name in data_name:
        if name not in name_list:
            noclass_data.append(name)
    for j in range(0, len(name_list)):
        name = name_list[j]
        duration = duration_list[j]
        if duration < min_class:
            noclass_data.append(name)
    return (noclass_data)

def get_noclass(data_name, data_class):  #获取全部考勤表缺勤数据
    data_noclass = []
    for data_list in data_class:
        class_date = data_list[0]
        class_subject = data_list[1]
        class_data = data_list[2]
        if len(class_data) > min_num:
            noclass_data = get_noclass_data(data_name, class_data)
            data_noclass.append([class_date, class_subject, noclass_data])
    return (data_noclass)

def get_noclass_times_data(data_name):  #获取单个科目缺勤次数
    noclass_times_data = []
    name_list = []
    times_list = []
    for name in data_name:
        if name not in name_list:
            name_list.append(name)
            times_list.append(1)
        else:
            p = name_list.index(name)
            times_list[p] += 1
    for k in range(0, len(name_list)):
        noclass_times_data.append([name_list[k], times_list[k]])
    return (noclass_times_data)

def get_noclass_times(data_noclass):  #获取全部科目缺勤次数
    global subject_list
    data_noclass_times = []
    subject_list = []
    name_list = []
    for data in data_noclass:
        subject = data[1]
        if subject not in subject_list:
            subject_list.append(subject)
    for data_subject in subject_list:
        data_name = []
        for data in data_noclass:
            subject = data[1]
            name_list = data[2]
            if subject == data_subject:
                for name in name_list:
                    data_name.append(name)
        noclass_times_data = get_noclass_times_data(data_name)
        data_noclass_times.append([data_subject, noclass_times_data])
    return (data_noclass_times)

def get_date_range(data_excel):  #获取考勤表日期范围
    date_list = []
    for data in data_excel:
        date = int(re.sub('-', '', data[0][:10]))
        date_list.append(date)
    return (str(min(date_list)) + '-' + str(max(date_list)))

def copy_excel(excel_name):  #复制excel表(我也不知道为什么不能直接复制)
    wb1 = openpyxl.load_workbook(class_path)
    ws1 = wb1[class_name]
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = class_name
    for i in range(1, class_y):
        wb1_list = []
        for cell in list(ws1.columns)[i]:
            wb1_list.append(cell.value)
        for j in range(len(wb1_list)):
            ws2.cell(j + 1, i, wb1_list[j])
    wb2.save(filename=excel_name)

def output_data(excel_name, data_name, data_noclass_times):  #生成考勤数据
    wb1 = openpyxl.load_workbook(excel_name)
    ws1 = wb1[class_name]
    for i in range(class_y, class_y + len(data_noclass_times)):
        subject = data_noclass_times[i - class_y][0]
        ws1.cell(1, i, subject)
        for j in range(2, len(data_name) + 2):
            name = data_name[j - 2]
            for k in range(0, len(data_noclass_times[i - class_y][1])):
                noclass_name = data_noclass_times[i - class_y][1][k][0]
                if name == noclass_name:
                    times = data_noclass_times[i - class_y][1][k][1]
                    ws1.cell(j, i, times)
    wb1.save(filename=excel_name)

def main():  #主函数
    data_excel = get_excel()  #考勤表信息(返回:[考勤表日期，考勤表路径，考勤表科目])
    #print(data_excel)
    data_name = get_name()  #本班名单(返回:[姓名])
    #print(data_name)
    data_class = get_class(data_excel, data_name)  #全部考勤表数据(返回:[日期，科目，考勤数据])
    #print(data_class)
    data_noclass = get_noclass(data_name, data_class)  #全部缺勤数据(返回:[日期，科目，缺勤数据])
    #print(data_noclass)
    data_noclass_times = get_noclass_times(data_noclass)  #科目缺勤数据(返回:[科目[姓名，缺勤次数]])
    #print(data_noclass_times)
    date_range = get_date_range(data_excel)  #考勤表日期范围(返回:开始日期-结束日期)
    #print(date_range)
    excel_name = date_range + class_name + '考勤数据.xlsx'  #考勤数据名称(返回:excel名称)
    copy_excel(excel_name)
    output_data(excel_name, data_name, data_noclass_times)  #输出考勤数据
    print('已生成' + excel_name)

def set_config():
    try:
        original_config = {"name_x": 1, "name_y": 3, "txkt_start_x": 5, "txkt_duration_y": 7, "txkt_name_y": 3, "class_y": 4,
                           "data_path": "./data", "class_path": "./class.xlsx", "min_class": 20, "min_num": 10}
        with open("./config.json", "w", encoding='utf-8') as f:
            json.dump(original_config, f, indent=4)
        print('生成配置文件成功!')
    except:
        print('生成配置文件出错!')
        time.sleep(3)
        exit()

if __name__ == '__main__':  #程序入口
    if not os.path.exists('./config.json'):
        print('未检测到配置文件，生成默认配置...')
        set_config()
        config = read_config()
    else:
        config = read_config()
    if config == True:
        if not os.path.exists(data_path):
            os.mkdir(data_path)
            print('未检测到考勤表，请将考勤表放置于' + data_path)
        class_name = input('请输入班级(班级名单中的表名称)：')
        try:
            main()
            time.sleep(3)
            exit()
        except:
            print('生成考勤数据失败!')